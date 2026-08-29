"""Origenes de archivo plano: Excel y CSV.

Sirven para dos cosas distintas. Una es operativa: hay empresas que exportan la
cartera del ERP a un archivo y no tienen API. La otra es de desarrollo: permiten
correr el motor contra el archivo de prueba y reconciliar el resultado, que es
el entregable verificable de la etapa 1 segun §8.
"""

from __future__ import annotations

import csv
from datetime import date, datetime
from pathlib import Path
from typing import Any, Iterator, Mapping, Sequence

from ..motores.cartera.datos import Movimiento
from .base import ErrorDeOrigen, MapeoCampos, normalizar

#: Mapeo del archivo de cartera de BUSINT.
#:
#: Es el contrato de datos de la etapa 0 para el origen plano. C-01 se resuelve
#: aqui: "Fecha Inicial" es la emision y "Fecha Final" el vencimiento, y es
#: contra esta ultima que se cuentan los dias.
MAPEO_BUSINT = MapeoCampos(
    cliente_nit="Nit Cliente",
    factura="Num Fact",
    fecha_emision="Fecha Inicial",
    fecha_vencimiento="Fecha Final",
    saldo="Valor Total",
    cliente_nombre="Razon Social",
    valor_credito="Valor Credito",
    vendedor="Vendedor",
    zona="Zona",
    ciudad="Ciudades",
    contacto="Celular",
    fecha_corte="Fecha de Corte",
    empresa_id=None,
    empresa_id_fijo=None,
)


class FuenteExcel:
    """Lee las cuentas abiertas de un archivo .xlsx.

    Requiere `openpyxl`, que es dependencia opcional: se importa dentro del
    metodo para que instalar el motor no obligue a instalar Excel.
    """

    def __init__(
        self,
        ruta: str | Path,
        mapeo: MapeoCampos = MAPEO_BUSINT,
        hoja: str | None = None,
        fila_encabezado: int = 1,
    ) -> None:
        self.ruta = Path(ruta)
        self.mapeo = mapeo
        self.hoja = hoja
        self.fila_encabezado = fila_encabezado

    def leer(self, empresa_id: str, corte: date) -> Iterator[Movimiento]:
        registros = list(self._registros())
        _verificar_corte(registros, self.mapeo, corte, str(self.ruta))
        mapeo = _con_empresa(self.mapeo, empresa_id)
        return normalizar(registros, mapeo, empresa_id)

    def _registros(self) -> Iterator[Mapping[str, Any]]:
        try:
            import openpyxl
        except ImportError:
            raise ErrorDeOrigen(
                "Leer Excel requiere openpyxl. Instala con: pip install 'busint-alertas[planos]'"
            ) from None

        if not self.ruta.exists():
            raise ErrorDeOrigen(f"No existe el archivo {self.ruta}")

        libro = openpyxl.load_workbook(self.ruta, data_only=True, read_only=True)
        try:
            hoja = libro[self.hoja] if self.hoja else libro[libro.sheetnames[0]]
            filas = hoja.iter_rows(values_only=True)
            for _ in range(self.fila_encabezado - 1):
                next(filas, None)
            encabezado = next(filas, None)
            if encabezado is None:
                raise ErrorDeOrigen(f"La hoja de {self.ruta} esta vacia.")
            columnas = [str(c).strip() if c is not None else "" for c in encabezado]
            for fila in filas:
                if all(v is None for v in fila):
                    continue
                yield dict(zip(columnas, fila))
        finally:
            libro.close()


class FuenteCSV:
    """Lee las cuentas abiertas de un CSV o de un archivo delimitado."""

    def __init__(
        self,
        ruta: str | Path,
        mapeo: MapeoCampos = MAPEO_BUSINT,
        delimitador: str = ",",
        codificacion: str = "utf-8-sig",
    ) -> None:
        self.ruta = Path(ruta)
        self.mapeo = mapeo
        self.delimitador = delimitador
        self.codificacion = codificacion

    def leer(self, empresa_id: str, corte: date) -> Iterator[Movimiento]:
        if not self.ruta.exists():
            raise ErrorDeOrigen(f"No existe el archivo {self.ruta}")
        with self.ruta.open(encoding=self.codificacion, newline="") as f:
            registros = [
                {(k.strip() if k else ""): v for k, v in fila.items()}
                for fila in csv.DictReader(f, delimiter=self.delimitador)
            ]
        _verificar_corte(registros, self.mapeo, corte, str(self.ruta))
        mapeo = _con_empresa(self.mapeo, empresa_id)
        yield from normalizar(registros, mapeo, empresa_id)


# --------------------------------------------------------------------------


def _con_empresa(mapeo: MapeoCampos, empresa_id: str) -> MapeoCampos:
    """C-08: si el archivo no trae empresa, se inyecta la que se esta procesando.

    Es un apano explicito y no un descuido: mientras el ERP no exponga la
    columna, un archivo plano pertenece por definicion a la empresa que lo
    exporto. Cuando la exponga, se nombra en el mapeo y esto deja de aplicar.
    """
    from dataclasses import replace

    if mapeo.empresa_id is None and mapeo.empresa_id_fijo is None:
        return replace(mapeo, empresa_id_fijo=empresa_id)
    return mapeo


def _verificar_corte(
    registros: Sequence[Mapping[str, Any]],
    mapeo: MapeoCampos,
    corte: date,
    origen: str,
) -> None:
    """Avisa si el archivo corresponde a un corte distinto del solicitado.

    Es un error facil de cometer y caro de detectar: se reprocesa el archivo del
    mes pasado y todos los indicadores salen mal sin que nada falle.
    """
    if mapeo.fecha_corte is None or not registros:
        return
    cortes = set()
    for r in registros:
        valor = r.get(mapeo.fecha_corte)
        if valor in (None, ""):
            continue
        if isinstance(valor, datetime):
            cortes.add(valor.date())
        elif isinstance(valor, date):
            cortes.add(valor)
        else:
            try:
                cortes.add(datetime.fromisoformat(str(valor).strip()).date())
            except ValueError:
                return
    if cortes and corte not in cortes:
        encontrados = ", ".join(str(c) for c in sorted(cortes))
        raise ErrorDeOrigen(
            f"{origen} corresponde al corte {encontrados}, pero se pidio {corte}. "
            f"Procesar un archivo de otro corte produce indicadores incorrectos."
        )
