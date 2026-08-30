"""Exportacion a Excel (§9).

"Todas las columnas originales mas codigo alerta, etiqueta, prioridad, accion,
estado, los dias calculados y el origen del dato de dias." El proposito no es
que quede bonito: es que alguien pueda auditar por que el sistema clasifico cada
factura como lo hizo, sin abrir el codigo.

Por eso van tambien el saldo bruto y el credito aplicado. Si el saldo no cuadra
con el ERP, la diferencia tiene que poder explicarse en la propia hoja (C-10).
"""

from __future__ import annotations

from datetime import date
from decimal import Decimal
from io import BytesIO
from typing import Any

from ..core.dinero import porcentaje
from .datos import Corte

CERO = Decimal("0.00")

#: Columnas de la hoja de alertas. El orden importa: primero identifican la
#: factura, luego la clasifican, y al final explican como se llego a ella.
COLUMNAS = (
    ("Nit Cliente", "cliente_nit"),
    ("Razon Social", "cliente_nombre"),
    ("Num Fact", "factura"),
    ("Dias Vencimiento", "dias"),
    ("Saldo", "saldo"),
    ("Saldo Bruto", "saldo_bruto"),
    ("Valor Credito Aplicado", "credito_aplicado"),
    ("Bucket", "bucket"),
    ("Codigo Alerta", "codigo"),
    ("Etiqueta Alerta", "etiqueta"),
    ("Prioridad", "prioridad_etiqueta"),
    ("Accion Sugerida", "accion"),
    ("Estado Alerta", "estado"),
    ("Origen Dias", "origen_dias"),
    ("Explicacion", "explicacion"),
    ("Vendedor", "vendedor"),
    ("Zona", "zona"),
)

FORMATO_PESOS = '"$" #,##0.00'


def generar(corte: Corte) -> bytes:
    """Devuelve el .xlsx del corte, en memoria."""
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    libro = openpyxl.Workbook()

    _hoja_alertas(libro.active, corte, Font, PatternFill, Alignment, get_column_letter)
    _hoja_clientes(libro.create_sheet("Riesgo por cliente"), corte, Font, get_column_letter)
    _hoja_aging(libro.create_sheet("Aging"), corte, Font, get_column_letter)
    _hoja_parametros(libro.create_sheet("Parametros"), corte, Font)

    memoria = BytesIO()
    libro.save(memoria)
    return memoria.getvalue()


def _encabezar(hoja, titulos, Font, PatternFill, get_column_letter, anchos=None):
    hoja.append(list(titulos))
    relleno = PatternFill("solid", fgColor="2F6B9A") if PatternFill else None
    for i, _ in enumerate(titulos, start=1):
        celda = hoja.cell(row=1, column=i)
        celda.font = Font(bold=True, color="FFFFFF")
        if relleno:
            celda.fill = relleno
        hoja.column_dimensions[get_column_letter(i)].width = (
            (anchos or {}).get(i, 18)
        )
    # Congelar la cabecera: una hoja de 500.000 filas es inutilizable sin esto.
    hoja.freeze_panes = "A2"


def _hoja_alertas(hoja, corte: Corte, Font, PatternFill, Alignment, get_column_letter):
    hoja.title = "Alertas"
    _encabezar(
        hoja, [t for t, _ in COLUMNAS], Font, PatternFill, get_column_letter,
        anchos={2: 34, 15: 60},
    )

    for alerta in corte.alertas:
        datos = alerta.datos or {}
        fila: list[Any] = []
        for _, campo in COLUMNAS:
            if campo == "cliente_nombre":
                fila.append(corte.nombre_de(alerta.cliente_nit))
            elif campo == "prioridad_etiqueta":
                fila.append(corte.etiqueta_prioridad(alerta.prioridad))
            elif campo in ("vendedor", "zona"):
                fila.append(datos.get(campo, ""))
            else:
                fila.append(getattr(alerta, campo, None))
        hoja.append(fila)

    columnas_dinero = [
        i for i, (_, c) in enumerate(COLUMNAS, start=1)
        if c in ("saldo", "saldo_bruto", "credito_aplicado")
    ]
    for fila in hoja.iter_rows(min_row=2):
        for i in columnas_dinero:
            fila[i - 1].number_format = FORMATO_PESOS
    hoja.auto_filter.ref = hoja.dimensions


def _hoja_clientes(hoja, corte: Corte, Font, get_column_letter):
    titulos = (
        "Nit Cliente", "Razon Social", "Cartera Total", "Por Vencer", "Vence Hoy",
        "Vencida", "% Vencida", "Mayor a 90", "% Mayor a 90", "Mayor a 150",
        "Dias Maximos Vencidos", "N Facturas", "N Facturas Vencidas",
        "Prioridad", "Marcadores",
    )
    hoja.append(list(titulos))
    for i in range(1, len(titulos) + 1):
        hoja.cell(row=1, column=i).font = Font(bold=True)
        hoja.column_dimensions[get_column_letter(i)].width = 20
    hoja.freeze_panes = "A2"

    for c in corte.clientes:
        hoja.append([
            c.cliente_nit, c.cliente_nombre, c.cartera_total, c.por_vencer,
            c.vence_hoy, c.vencida, c.pct_vencida, c.mayor_90, c.pct_90,
            c.mayor_150, c.dias_max, c.n_facturas, c.n_vencidas,
            corte.etiqueta_prioridad(c.prioridad), ", ".join(c.marcadores or []),
        ])
    for fila in hoja.iter_rows(min_row=2):
        for i in (3, 4, 5, 6, 8, 10):
            fila[i - 1].number_format = FORMATO_PESOS


def _hoja_aging(hoja, corte: Corte, Font, get_column_letter):
    titulos = ("Bucket", "Etiqueta", "Desde", "Hasta", "Saldo", "Facturas", "% del Total")
    hoja.append(list(titulos))
    for i in range(1, len(titulos) + 1):
        hoja.cell(row=1, column=i).font = Font(bold=True)
        hoja.column_dimensions[get_column_letter(i)].width = 18

    total = corte.cartera_total
    for b in corte.buckets:
        saldo = corte.totales_por_bucket.get(b.codigo, CERO)
        hoja.append([
            b.codigo, b.etiqueta, b.desde, b.hasta, saldo,
            corte.facturas_por_bucket.get(b.codigo, 0), porcentaje(saldo, total),
        ])
    # La identidad de C-14, explicita tambien en la exportacion.
    hoja.append([])
    hoja.append(["", "Cartera total", "", "", corte.cartera_total])
    hoja.append(["", "= Por vencer", "", "", corte.por_vencer])
    hoja.append(["", "+ Vence hoy", "", "", corte.vence_hoy])
    hoja.append(["", "+ Vencida", "", "", corte.vencida])
    for fila in hoja.iter_rows(min_row=2):
        fila[4].number_format = FORMATO_PESOS


def _hoja_parametros(hoja, corte: Corte, Font):
    """Con que configuracion se calculo este corte.

    Sin esto, dos exportaciones del mismo corte con umbrales distintos serian
    indistinguibles y la auditoria dejaria de ser posible.
    """
    hoja.append(["Campo", "Valor"])
    hoja.cell(row=1, column=1).font = Font(bold=True)
    hoja.cell(row=1, column=2).font = Font(bold=True)
    hoja.column_dimensions["A"].width = 28
    hoja.column_dimensions["B"].width = 44

    for campo, valor in (
        ("Empresa", corte.empresa_id),
        ("Fecha de corte", corte.corte),
        ("Generado", corte.generado),
        ("Version de parametros", corte.version_parametros),
        ("Clientes", len(corte.clientes)),
        ("Alertas", len(corte.alertas)),
    ):
        hoja.append([campo, valor])
